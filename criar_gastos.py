import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def criar_arquivo_gastos():
    """Cria o arquivo gastos.xlsx com estrutura para registrar gastos diários."""
    
    if os.path.exists('gastos.xlsx'):
        print("Arquivo gastos.xlsx já existe.")
        return
    
    # Cria um novo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"
    
    # Define cabeçalhos
    ws['A1'] = "Data"
    ws['B1'] = "Descrição"
    ws['C1'] = "Valor"
    
    # Formata cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in ['A1', 'B1', 'C1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal="center")
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    
    # Salva o arquivo
    wb.save('gastos.xlsx')
    print("Arquivo gastos.xlsx criado com sucesso!")

if __name__ == "__main__":
    criar_arquivo_gastos()
